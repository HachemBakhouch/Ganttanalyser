import os
import openpyxl
import csv
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from typing import List, Dict, Any
from src.data.data_models import ProfileWorkload, WorkloadEntry, ExportConfiguration


class ExportService:
    """
    Service responsable de l'exportation des résultats d'analyse
    """

    def __init__(self):
        self.styles = getSampleStyleSheet()

    def export_txt(
        self,
        file_path: str,
        profiles_workload: List[ProfileWorkload],
        detailed_workload: Dict[str, Dict[str, List[WorkloadEntry]]],
    ):
        """
        Exporte les résultats au format texte

        :param file_path: Chemin du fichier de sortie
        :param profiles_workload: Charge de travail globale par profil
        :param detailed_workload: Charge de travail détaillée
        """
        with open(file_path, "w", encoding="utf-8") as f:
            # Résultats globaux
            f.write("RÉSULTATS GLOBAUX PAR PROFIL:\n")
            f.write("==========================\n\n")
            for profile in profiles_workload:
                f.write(f"Profil: {profile.profile}\n")
                f.write(
                    f"Charge de travail totale: {profile.total_workload:.2f} heures\n\n"
                )

            # Résultats détaillés
            f.write("RÉSULTATS DÉTAILLÉS PAR CHEF DE PROJET ET PAR PROJET:\n")
            f.write("=================================================\n\n")

            for pm, projects in detailed_workload.items():
                f.write(f"Chef de projet: {pm}\n")
                f.write("-" * 50 + "\n")

                for project, entries in projects.items():
                    f.write(f"  Projet: {project}\n")
                    for entry in entries:
                        jira_info = (
                            f" (JIRA: {entry.jira_ticket})" if entry.jira_ticket else ""
                        )
                        f.write(
                            f"    • {entry.profile}: {entry.workload:.2f} heures{jira_info}\n"
                        )
                    f.write("\n")

    def export_xlsx(
        self,
        file_path: str,
        profiles_workload: List[ProfileWorkload],
        detailed_workload: Dict[str, Dict[str, List[WorkloadEntry]]],
    ):
        """
        Exporte les résultats au format Excel

        :param file_path: Chemin du fichier de sortie
        :param profiles_workload: Charge de travail globale par profil
        :param detailed_workload: Charge de travail détaillée
        """
        wb = openpyxl.Workbook()

        # Feuille des résultats globaux
        ws_global = wb.active
        ws_global.title = "Résultats Globaux"

        ws_global.cell(row=1, column=1, value="Profil")
        ws_global.cell(row=1, column=2, value="Charge de travail totale (heures)")

        for idx, profile in enumerate(profiles_workload, start=2):
            ws_global.cell(row=idx, column=1, value=profile.profile)
            ws_global.cell(row=idx, column=2, value=profile.total_workload)

        # Feuille des résultats détaillés
        ws_detailed = wb.create_sheet(title="Résultats Détaillés")

        ws_detailed.cell(row=1, column=1, value="Chef de projet")
        ws_detailed.cell(row=1, column=2, value="Projet")
        ws_detailed.cell(row=1, column=3, value="Profil")
        ws_detailed.cell(row=1, column=4, value="Charge (heures)")
        ws_detailed.cell(row=1, column=5, value="Ticket JIRA")

        row_idx = 2
        for pm, projects in detailed_workload.items():
            for project, entries in projects.items():
                for entry in entries:
                    ws_detailed.cell(row=row_idx, column=1, value=pm)
                    ws_detailed.cell(row=row_idx, column=2, value=project)
                    ws_detailed.cell(row=row_idx, column=3, value=entry.profile)
                    ws_detailed.cell(row=row_idx, column=4, value=entry.workload)
                    ws_detailed.cell(
                        row=row_idx, column=5, value=entry.jira_ticket or "N/A"
                    )
                    row_idx += 1

        wb.save(file_path)

    def export_pdf(
        self,
        file_path: str,
        profiles_workload: List[ProfileWorkload],
        detailed_workload: Dict[str, Dict[str, List[WorkloadEntry]]],
    ):
        """
        Exporte les résultats au format PDF

        :param file_path: Chemin du fichier de sortie
        :param profiles_workload: Charge de travail globale par profil
        :param detailed_workload: Charge de travail détaillée
        """
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        elements = []

        # Styles personnalisés
        title_style = self.styles["Heading1"]
        subtitle_style = self.styles["Heading2"]
        normal_style = self.styles["Normal"]

        # Titre
        elements.append(Paragraph("Analyse de Charge de Travail", title_style))
        elements.append(Spacer(1, 12))

        # Résultats globaux
        elements.append(Paragraph("Résultats Globaux par Profil", subtitle_style))
        for profile in profiles_workload:
            elements.append(
                Paragraph(
                    f"Profil: {profile.profile} - Charge totale: {profile.total_workload:.2f} heures",
                    normal_style,
                )
            )

        elements.append(Spacer(1, 12))

        # Résultats détaillés
        elements.append(
            Paragraph("Résultats Détaillés par Chef de Projet", subtitle_style)
        )

        for pm, projects in detailed_workload.items():
            elements.append(Paragraph(f"Chef de projet: {pm}", subtitle_style))

            for project, entries in projects.items():
                elements.append(Paragraph(f"Projet: {project}", normal_style))

                # Créer un tableau pour les entrées de ce projet
                table_data = [["Profil", "Charge (heures)", "Ticket JIRA"]]
                for entry in entries:
                    table_data.append(
                        [
                            entry.profile,
                            f"{entry.workload:.2f}",
                            entry.jira_ticket or "N/A",
                        ]
                    )

                table = Table(table_data)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 12),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )
                elements.append(table)
                elements.append(Spacer(1, 12))

        doc.build(elements)

    def export(
        self,
        config: ExportConfiguration,
        profiles_workload: List[ProfileWorkload],
        detailed_workload: Dict[str, Dict[str, List[WorkloadEntry]]],
    ):
        """
        Exporte les résultats selon la configuration

        :param config: Configuration d'exportation
        :param profiles_workload: Charge de travail globale par profil
        :param detailed_workload: Charge de travail détaillée
        """
        if not config.file_path:
            raise ValueError("Le chemin du fichier n'est pas spécifié")

        if config.export_format == "txt":
            self.export_txt(config.file_path, profiles_workload, detailed_workload)
        elif config.export_format == "xlsx":
            self.export_xlsx(config.file_path, profiles_workload, detailed_workload)
        elif config.export_format == "pdf":
            self.export_pdf(config.file_path, profiles_workload, detailed_workload)
        else:
            raise ValueError(
                f"Format d'exportation non supporté: {config.export_format}"
            )
