import openpyxl
from openpyxl.utils import column_index_from_string
from typing import List, Dict, Any, Optional
from src.data.data_models import WorkloadEntry, AnalysisConfiguration


class ExcelReader:
    def __init__(self, file_path: str):
        """
        Initialise le lecteur de fichier Excel

        :param file_path: Chemin du fichier Excel à charger
        """
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        self._load_workbook()

    def _load_workbook(self):
        """
        Charge le fichier Excel
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            self.sheet = self.workbook.active
        except Exception as e:
            raise ValueError(f"Impossible de charger le fichier Excel: {str(e)}")

    def extract_unique_profiles(self, config: AnalysisConfiguration) -> List[str]:
        """
        Extrait les profils uniques du fichier Excel

        :param config: Configuration pour l'extraction
        :return: Liste des profils uniques
        """
        unique_profiles = set()
        profile_col_idx = column_index_from_string(config.profile_column)

        for row in range(config.start_row, config.end_row + 1):
            cell_value = self.sheet.cell(row=row, column=profile_col_idx).value
            if cell_value:
                unique_profiles.add(str(cell_value))

        return list(unique_profiles)

    def read_workload_entries(
        self, config: AnalysisConfiguration
    ) -> List[WorkloadEntry]:
        """
        Lit les entrées de charge de travail du fichier Excel

        :param config: Configuration pour la lecture
        :return: Liste des entrées de charge de travail
        """
        workload_entries = []

        # Indices des colonnes
        profile_col_idx = column_index_from_string(config.profile_column)
        start_col_idx = column_index_from_string(config.start_column)
        end_col_idx = column_index_from_string(config.end_column)
        pm_col_idx = column_index_from_string("B")  # Colonne chef de projet
        project_col_idx = column_index_from_string("D")  # Colonne projet
        jira_col_idx = column_index_from_string("F")  # Colonne ticket JIRA

        for row in range(config.start_row, config.end_row + 1):
            # Vérifier si le profil est dans la liste des profils sélectionnés (si fournie)
            profile = self.sheet.cell(row=row, column=profile_col_idx).value

            if not config.selected_profiles or (
                profile and str(profile) in config.selected_profiles
            ):

                # Calculer la charge de travail totale pour cette ligne
                row_total = sum(
                    cell.value or 0
                    for col_idx in range(start_col_idx, end_col_idx + 1)
                    for cell in [self.sheet.cell(row=row, column=col_idx)]
                    if isinstance(cell.value, (int, float))
                )

                # Récupérer les autres informations
                pm_name = self.sheet.cell(row=row, column=pm_col_idx).value
                project_name = self.sheet.cell(row=row, column=project_col_idx).value
                jira_ticket = self.sheet.cell(row=row, column=jira_col_idx).value

                if pm_name and project_name:
                    entry = WorkloadEntry(
                        project_manager=str(pm_name),
                        project=str(project_name),
                        profile=str(profile),
                        jira_ticket=str(jira_ticket) if jira_ticket else None,
                        workload=row_total,
                    )
                    workload_entries.append(entry)

        return workload_entries

    def get_sheet_names(self) -> List[str]:
        """
        Retourne les noms des feuilles dans le classeur

        :return: Liste des noms de feuilles
        """
        return self.workbook.sheetnames if self.workbook else []
